# pandas example table
import numpy as np
import pandas as pd

data = pd.DataFrame(
    np.random.randn(10, 4),
    columns=list("ABCD"),
)

here = "office"

# make an excel with the table
data.to_excel(f"{here}/table.xlsx", index=True)

# conditional formatting on existing excel
################################################
import openpyxl

wb = openpyxl.load_workbook(f"{here}/table.xlsx")

ws = wb.active

red_text = openpyxl.styles.Font(color="00FF0000")

for row in ws.iter_rows(min_row=2, min_col=2, max_col=5, max_row=11):
    for cell in row:
        if cell.value < 0:
            cell.font = red_text

wb.save(f"{here}/table.xlsx")

# conditional formatting on a dataframe then save to excel
########################################################

# work thru the dataframe, create cells with the color formatting

wb = openpyxl.Workbook()
ws = wb.active

for i in range(data.shape[0]):
    for j in range(data.shape[1]):
        cell = ws.cell(row=i + 2, column=j + 2)
        if data.iloc[i, j] < 0:
            cell.font = red_text
        cell.value = data.iloc[i, j]

wb.save(f"{here}/table2.xlsx")


# use color scales to show the range of values, red to green
############################################################

wb = openpyxl.Workbook()
ws = wb.active

for i in range(data.shape[0]):
    for j in range(data.shape[1]):
        cell = ws.cell(row=i + 2, column=j + 2)
        cell.value = data.iloc[i, j]

ws.conditional_formatting.add(
    f"B2:E{data.shape[0]+1}",
    openpyxl.formatting.rule.ColorScaleRule(
        start_type="min",
        start_color="00FF0000",
        end_type="max",
        end_color="0000FF00",
    ),
)

wb.save(f"{here}/table3.xlsx")
